#%%
import os
import json
import smtplib
import pickle
import base64
import pandas as pd
import plotly.express as px
import streamlit as st
import yaml
import io

from datetime import datetime
from pathlib import Path
from yaml.loader import SafeLoader
from PIL import Image

from email import encoders
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border

import streamlit_authenticator as stauth
from teacher_alocation import TeacherScheduler


def replicate_row(row: pd.Series, times: int) -> pd.DataFrame:
    hora_inicial = pd.to_datetime(row['horario'], format='%H:%M:%S')
    novas_linhas = []
    for i in range(times):
        nova_linha = row.copy()
        nova_linha['horario'] = (hora_inicial + pd.Timedelta(hours=i)).strftime('%H:%M:%S')
        novas_linhas.append(nova_linha)
    return pd.DataFrame(novas_linhas)

def desaninhar_dias(df):
    df['dias da semana'] = df['dias da semana'].str.split(',')
    df = df.explode('dias da semana')
    return df

def expand_rows( df: pd.DataFrame, func) -> pd.DataFrame:
    return pd.concat(df.apply(func, axis=1).tolist(), ignore_index=True)

def clean_data(df: pd.DataFrame)-> pd.DataFrame:

    df['intenviso'] = np.where(df['n aulas']>=10,1,0)

    df['dias da semana'] = df['dias da semana'].str.replace('●',',').str.replace(' ','').str.replace('DOUBLE',',').str.replace('-Triple','').str.split(',')
    df = df.explode('dias da semana').reset_index(drop=True)

    df['dias da semana'] = df['dias da semana'].str.replace('ª','ª,').str.split(',')
    df = df.explode('dias da semana').reset_index(drop=True)

    df = df[df['dias da semana']!=''].copy()

    substituicoes = {
        '2ª': 'SEGUNDA',
        '3ª': 'TERÇA',
        '4ª': 'QUARTA',
        '5ª': 'QUINTA',
        '6ª': 'SEXTA',
        'Saturday': 'SÁBADO'
    }

    df['dias da semana'] = df['dias da semana'].replace(substituicoes, regex=True)

    df['status'].fillna('PRESENCIAL', inplace=True)

    df['horario_tratado'] = pd.to_datetime(df['horario'], format='%H:%M:%S')

    return df

def filter_class_without_teacher(df: pd.DataFrame) -> pd.DataFrame:
    return df[((df['teacher'].isnull()) | (df['teacher']=='-'))]

def base_selection(df: pd.DataFrame) -> tuple:
    aulas_tratadas = df.loc[~df['nome grupo'].isnull()]

    aulas = aulas_tratadas.copy()
    # tratando os dados para colocar cada linha uma aula
    aulas['dias da semana'] = aulas['dias da semana'].str.replace('EVERYDAY','2ª ● 3ª ● 4ª ● 5ª ● 6ª')
    aulas['stage'] = aulas['stage'].apply(lambda x: f'ESTAGIO_{x}' if isinstance(x, int) else x)
    aulas['ultimo_professor'] = aulas['ultimo_professor'].astype(str)
    aulas['penultimo_professor'] = aulas['penultimo_professor'].astype(str)


    # separando as aulas que são triplas, duplas e o resto
    tri = aulas.loc[aulas['dias da semana'] == 'Saturday - Triple']
    doub = aulas.loc[aulas['dias da semana'].str.contains('DOUBLE')]
    aulas_simples = aulas[~aulas['dias da semana'].str.contains('DOUBLE|Saturday - Triple')].copy()

    # tratando a colunas horario
    aulas_simples['horario'] = pd.to_datetime(aulas_simples['horario'],format='%H:%M:%S').dt.strftime('%H:%M:%S')
    return aulas_simples, doub, tri

def transform_classes_dateframe(aulas_raw):
    #aulas_filtrada = filter_class_without_teacher(aulas_raw)
    aulas_simples,doub,tri=base_selection(aulas_raw)

    # transformando aulas duplas/triplas em 2/3 linhas
    aulas_duplicadas = expand_rows(doub, lambda row: replicate_row(row, times=2))
    aulas_triplicadas = expand_rows(tri, lambda row: replicate_row(row, times=3))

    # juntando todas as linhas
    df_tratado = pd.concat([aulas_simples, aulas_duplicadas, aulas_triplicadas], ignore_index=True)

    df_result = clean_data(df_tratado)

    return df_result

def transform_teacher_dataframe(professores_raw):
    professores_raw.columns = professores_raw.columns.astype(str)
    return professores_raw

def transform_alocation_dataframe(aulas_raw,base_alocada):
    alocation_df = pd.merge(aulas_raw, base_alocada, on='nome grupo', how='left')
    alocation_df['penultimo_professor'] = alocation_df['ultimo_professor']
    alocation_df.loc[alocation_df['professores_alocados'].notnull(), 'teacher'] = alocation_df.loc[alocation_df['professores_alocados'].notnull(), 'professores_alocados']
    alocation_df['ultimo_professor'] = alocation_df['teacher']
    alocation_df.drop(columns=['professores_alocados'], inplace=True)
    not_alocation_df = alocation_df.loc[((alocation_df['teacher'].isnull())|(alocation_df['teacher']=='-'))].copy()
    return alocation_df, not_alocation_df

def enviar_email_para_todos(combined_df, arquivo_original):
    log_messages = []
    failed_teachers = []

    wb_original = load_workbook(arquivo_original)
    ws_original = wb_original.active

    for teacher in combined_df['Teacher'].unique():
        professor_data = combined_df[combined_df['Teacher'] == teacher]

        if professor_data.empty or pd.isna(teacher) or teacher == '':
            continue

        nome_grupo = professor_data['Nome Grupo'].to_list()
        email_professor = professor_data['Email'].values[0]

        wb_professor = load_workbook(arquivo_original)
        ws_professor = wb_professor.active

        # Ocultar as linhas que não correspondem ao professor
        for row_idx in range(2, ws_professor.max_row + 1):
            nome_grupo_linha = ws_professor.cell(row=row_idx, column=1).value
            if nome_grupo_linha not in nome_grupo:
                ws_professor.row_dimensions[row_idx].hidden = True
            else:
                ws_professor.row_dimensions[row_idx].hidden = False

        # Salvar o arquivo filtrado para o professor
        diretorio_atual = os.path.dirname(os.path.abspath(__file__))
        diretorio_destino = os.path.join(diretorio_atual, "alocacoes_final")
        if not os.path.exists(diretorio_destino):
            os.makedirs(diretorio_destino)

        arquivo_nuvem = os.path.join(diretorio_destino, f"{teacher}_alocacao.xlsx")
        wb_professor.save(arquivo_nuvem)

        nome_grupo_formatado = "\n".join(f"    {grupo}" for grupo in nome_grupo)

        message = (
            f"Olá {teacher},\n\n"
            f"Você foi alocado nas seguintes turmas:\n\n"
            f"{nome_grupo_formatado}\n\n"
            f"Segue abaixo, o anexo com as informações detalhadas das turmas.\n\n"
            f"Atenciosamente,\n"
            f"The Family Idiomas"
        )

        with open('.devcontainer/config.json') as f:
            config = json.load(f)

        from_email = 'teacher.scheduler.contact@gmail.com'
        to_email = email_professor
        password = config["email_password"]

        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = 'Rota das Turmas'

        msg.attach(MIMEText(message, 'plain'))

        with open(arquivo_nuvem, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename={os.path.basename(arquivo_nuvem)}")
            msg.attach(part)

        try:
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(from_email, password)
            text = msg.as_string()
            server.sendmail(from_email, to_email, text)
            server.quit()
            log_messages.append(f"✅ E-mail enviado com sucesso para {teacher}")
        except Exception as e:
            log_messages.append(f"❌ Falha ao enviar e-mail para {teacher}")
            failed_teachers.append(teacher)

        os.remove(arquivo_nuvem)
        print(f'Arquivo {arquivo_nuvem} excluído com sucesso.')

    st.session_state.failed_teachers = failed_teachers
    return log_messages